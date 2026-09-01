"""Import the KCD-9 complete-code sheet into PostgreSQL."""

from __future__ import annotations

import argparse

from openpyxl import load_workbook
from sqlalchemy import delete

from app.database import SessionLocal, engine
from app.models.base import Base
from app.models.kcd_code import KcdCode


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", help="KCD-9 disease master XLSX path")
    args = parser.parse_args()

    workbook = load_workbook(args.xlsx, read_only=True, data_only=True)
    sheet = workbook["상병분류기호(완전코드)"]
    Base.metadata.create_all(bind=engine)

    inserted = 0
    batch: list[dict[str, str | None]] = []
    with SessionLocal() as db:
        db.execute(delete(KcdCode))
        for row in sheet.iter_rows(min_row=12, values_only=True):
            code, name_ko, name_en = row[1], row[2], row[3]
            if not code or not name_ko:
                continue
            batch.append(
                {
                    "code": str(code).replace(".", "").strip().upper(),
                    "name_ko": str(name_ko).strip(),
                    "name_en": str(name_en).strip() if name_en else None,
                }
            )
            if len(batch) >= 2000:
                db.execute(KcdCode.__table__.insert(), batch)
                inserted += len(batch)
                batch.clear()
        if batch:
            db.execute(KcdCode.__table__.insert(), batch)
            inserted += len(batch)
        db.commit()
    print(f"Imported {inserted} KCD-9 rows")


if __name__ == "__main__":
    main()
